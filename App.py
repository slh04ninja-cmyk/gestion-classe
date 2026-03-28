# ========================= PARTIE 1 =========================
import streamlit as st
import pandas as pd
import json
import io
from datetime import datetime
from supabase import create_client, Client

st.set_page_config(
    page_title="Gestion de Classe",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Chargement du CSS externe ──────────────────────────────────────────
with open("style.css", "r", encoding="utf-8") as f:
    css_content = f.read()
st.markdown(f"<style>{css_content}</style>", unsafe_allow_html=True)

# ── Vérification des secrets ────────────────────────────────────────────
try:
    SUPABASE_URL = st.secrets["supabase_url"]
    SUPABASE_KEY = st.secrets["supabase_key"]
except KeyError as e:
    st.error(f"❌ Secret manquant : {e}. Veuillez configurer les secrets dans Streamlit Cloud.")
    st.stop()

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Fonctions utilitaires ──────────────────────────────────────────────
def get_mention(score):
    if score >= 16:   return "Très Bien",  "m-tb"
    elif score >= 14: return "Bien",        "m-b"
    elif score >= 12: return "Assez Bien",  "m-ab"
    elif score >= 10: return "Passable",    "m-p"
    else:             return "Echec",       "m-ec"

def load_db():
    try:
        response = supabase.table("classes").select("*").execute()
        db = {}
        for row in response.data:
            db[row["id"]] = row["data"]
        return db
    except Exception as e:
        st.error(f"Erreur de chargement : {e}")
        return {}

def save_db(db):
    try:
        for class_id, class_data in db.items():
            existing = supabase.table("classes").select("id").eq("id", class_id).execute()
            if existing.data:
                supabase.table("classes").update({"data": class_data}).eq("id", class_id).execute()
            else:
                supabase.table("classes").insert({"id": class_id, "data": class_data}).execute()
        all_ids = {row["id"] for row in supabase.table("classes").select("id").execute().data}
        for class_id in all_ids - set(db.keys()):
            supabase.table("classes").delete().eq("id", class_id).execute()
    except Exception as e:
        st.error(f"Erreur de sauvegarde : {e}")
# ========================= PARTIE 2 =========================
if "db" not in st.session_state:
    st.session_state.db = load_db()
if "selected_class" not in st.session_state:
    st.session_state.selected_class = None
if "confirm_delete" not in st.session_state:
    st.session_state.confirm_delete = None

db = st.session_state.db

st.markdown("""
<div class="app-header">
    <div class="app-tag">Outil Pédagogique</div>
    <div class="app-title">📚 Gestion de Classe</div>
    <div class="app-sub">Suivi des notes de comportement — Accumulation +1 / -1</div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 📂 Mes Classes")
    if db:
        for cid, cls in db.items():
            label = cls["nom"] + " — " + cls["annee"]
            if st.button(label, key=f"sel_{cid}", use_container_width=True):
                st.session_state.selected_class = cid
                st.rerun()
    else:
        st.info("Aucune classe créée.")
    st.markdown("---")
    if st.button("➕ Nouvelle classe", use_container_width=True):
        st.session_state.selected_class = "new"
        st.rerun()

tab_home, tab_new, tab_class = st.tabs(["🏠 Accueil", "➕ Créer une classe", "📋 Gérer la classe"])

with tab_home:
    if not db:
        st.markdown("""
        <div style="text-align:center;padding:48px 20px;color:#546E6A">
            <div style="font-size:56px;margin-bottom:16px">📚</div>
            <div style="font-size:18px;font-weight:800;color:#00796B;margin-bottom:8px">
                Aucune classe enregistrée
            </div>
            <div style="font-size:14px">
                Créez votre première classe dans l'onglet <b>Créer une classe</b>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">📊 Vue d\'ensemble</div>', unsafe_allow_html=True)

        nb_classes   = len(db)
        nb_etudiants = sum(len(c["etudiants"]) for c in db.values())
        c1, c2 = st.columns(2)
        c1.metric("Classes enregistrées", nb_classes)
        c2.metric("Total étudiants",      nb_etudiants)
        st.markdown("---")

        for cid, cls in db.items():
            etudiants = cls["etudiants"]
            nb = len(etudiants)
            scores = [e["score"] for e in etudiants.values()]
            moy = sum(scores) / nb if nb else 0

            col_a, col_b, col_c = st.columns([3,1,1])
            col_a.markdown(
                '<b style="font-size:15px">' + cls["nom"] + '</b>'
                ' <span style="color:#546E6A;font-size:13px">— ' + cls["annee"] + '</span>',
                unsafe_allow_html=True
            )
            col_b.markdown(
                '<div style="text-align:center;font-size:13px;color:#546E6A">'
                + str(nb) + ' étudiants</div>',
                unsafe_allow_html=True
            )
            col_c.markdown(
                '<div style="text-align:center;font-size:13px;font-weight:800;color:#00796B">'
                'Moy: ' + "{:.1f}".format(moy) + '</div>',
                unsafe_allow_html=True
            )
        st.markdown('</div>', unsafe_allow_html=True)

with tab_new:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">➕ Nouvelle Classe</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        nom_classe  = st.text_input("Nom de la classe", placeholder="ex: 2BACSPF-1")
    with col2:
        annee_scol  = st.text_input("Année scolaire", placeholder="ex: 2025/2026")

    st.markdown("**Liste des étudiants**")
    methode = st.radio("Méthode de saisie", ["Saisie manuelle", "Import Excel/CSV"], horizontal=True)

    etudiants_input = []

    if methode == "Saisie manuelle":
        nb_students = st.number_input("Nombre d'étudiants", min_value=1, max_value=60, value=5, step=1)
        st.markdown("*Saisissez un nom par ligne :*")
        noms_text = st.text_area(
            "Noms des étudiants (un par ligne)",
            height=max(120, int(nb_students) * 22),
            placeholder="Ahmed Benali\nFatima Zahra\nYoussef Alami\n..."
        )
        if noms_text.strip():
            etudiants_input = [n.strip() for n in noms_text.strip().split("\n") if n.strip()]
    else:
        uploaded_list = st.file_uploader("Fichier Excel ou CSV (1 colonne = noms)", type=["xlsx","xls","csv"])
        if uploaded_list:
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    if uploaded_list.name.endswith(".csv"):
                        df_list = pd.read_csv(io.BytesIO(uploaded_list.getvalue()), header=None)
                    else:
                        df_list = pd.read_excel(io.BytesIO(uploaded_list.getvalue()), header=None, engine="openpyxl")
                for col in df_list.columns:
                    vals = df_list[col].dropna().astype(str).str.strip()
                    vals = vals[vals.str.len() > 1]
                    if len(vals) > 0:
                        etudiants_input = vals.tolist()
                        break
                st.success(f"✅ {len(etudiants_input)} étudiants détectés")
                st.write(etudiants_input[:5])
            except Exception as e:
                st.error(f"Erreur : {e}")

    if etudiants_input:
        st.info(f"👥 {len(etudiants_input)} étudiants prêts à être enregistrés")

    if st.button("💾 Créer la classe", type="primary"):
        if not nom_classe.strip():
            st.error("❌ Le nom de la classe est obligatoire.")
        elif not annee_scol.strip():
            st.error("❌ L'année scolaire est obligatoire.")
        elif not etudiants_input:
            st.error("❌ Ajoutez au moins un étudiant.")
        else:
            cid = nom_classe.strip().replace(" ", "_") + "_" + annee_scol.strip().replace("/", "-")
            if cid in db:
                st.error("❌ Une classe avec ce nom et cette année existe déjà.")
            else:
                db[cid] = {
                    "nom":       nom_classe.strip(),
                    "annee":     annee_scol.strip(),
                    "created":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "etudiants": {
                        str(i): {
                            "nom":       nom,
                            "score":     0,
                            "historique":[]
                        }
                        for i, nom in enumerate(etudiants_input)
                    }
                }
                save_db(db)
                st.session_state.selected_class = cid
                st.success(f"✅ Classe '{nom_classe}' créée avec {len(etudiants_input)} étudiants !")
                st.balloons()
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)
# ========================= PARTIE 3a =========================
with tab_class:
    if not db:
        st.info("Aucune classe disponible. Créez-en une d'abord.")
    else:
        classe_options = {cid: cls["nom"] + " — " + cls["annee"] for cid, cls in db.items()}
        selected = st.selectbox("Choisir une classe", options=list(classe_options.keys()),
                                format_func=lambda x: classe_options[x],
                                index=list(classe_options.keys()).index(st.session_state.selected_class)
                                if st.session_state.selected_class in classe_options else 0)
        st.session_state.selected_class = selected
        cls = db[selected]

        stab1, stab2, stab3, stab4 = st.tabs(["⚡ Saisie Notes", "📊 Classement", "📜 Historique", "⚙️ Gérer"])

        with stab1:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown(
                '<div class="card-title">⚡ Saisie rapide — '
                + cls["nom"] + ' &nbsp;<span style="font-size:13px;color:#546E6A">'
                + cls["annee"] + '</span></div>',
                unsafe_allow_html=True
            )

            etudiants = cls["etudiants"]
            nb = len(etudiants)
            scores = [e["score"] for e in etudiants.values()]
            moy = sum(scores) / nb if nb else 0

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Étudiants", nb)
            c2.metric("Moyenne", "{:.2f}".format(moy))
            c3.metric("Meilleur", max(scores) if scores else 0)
            c4.metric("Plus bas", min(scores) if scores else 0)
            st.markdown("---")

            # Note globale
            st.markdown("**Note globale (toute la classe)**")
            col_g1, col_g2, col_g3, col_g4 = st.columns([1,1,1,3])
            note_globale = None
            with col_g1:
                if st.button("➕ +1 Classe", key="g_plus"):
                    note_globale = 1
            with col_g2:
                if st.button("➖ -1 Classe", key="g_moins"):
                    note_globale = -1
            with col_g3:
                # Placeholder pour alignement
                pass
            with col_g4:
                motif_global = st.text_input("Motif (optionnel)", key="motif_global",
                                             placeholder="ex: Participation active, Retard...")

            if note_globale is not None:
                ts = datetime.now().strftime("%d/%m/%Y %H:%M")
                for eid in etudiants:
                    etudiants[eid]["score"] += note_globale
                    etudiants[eid]["historique"].append({
                        "note": note_globale, "motif": motif_global,
                        "date": ts, "type": "global"
                    })
                save_db(db)
                st.success(f"✅ Note {'+1' if note_globale==1 else '-1'} appliquée à toute la classe !")
                st.rerun()

            st.markdown("---")
            st.markdown("**Notes individuelles**")

            # Affichage des étudiants avec carte simplifiée
            for eid, etudiant in sorted(etudiants.items(), key=lambda x: x[1]["nom"]):
                score = etudiant["score"]
                mention, m_cls = get_mention(score)

                # Carte en flex
                st.markdown('<div class="student-card">', unsafe_allow_html=True)
                st.markdown('<div class="student-info">', unsafe_allow_html=True)
                st.markdown(f'<span class="student-name">{etudiant["nom"]}</span>', unsafe_allow_html=True)
                st.markdown(f'<span class="student-score">{score}</span>', unsafe_allow_html=True)
                st.markdown(f'<span class="mention-badge {m_cls}">{mention}</span>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

                # Boutons +1 et -1 côte à côte
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("➕ +1", key=f"p_{eid}", use_container_width=True):
                        ts = datetime.now().strftime("%d/%m/%Y %H:%M")
                        etudiants[eid]["score"] += 1
                        etudiants[eid]["historique"].append({"note":1, "motif": "", "date": ts, "type": "indiv"})
                        save_db(db)
                        st.rerun()
                with col_btn2:
                    if st.button("➖ -1", key=f"m_{eid}", use_container_width=True):
                        ts = datetime.now().strftime("%d/%m/%Y %H:%M")
                        etudiants[eid]["score"] -= 1
                        etudiants[eid]["historique"].append({"note":-1, "motif": "", "date": ts, "type": "indiv"})
                        save_db(db)
                        st.rerun()

                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)
# ========================= PARTIE 3b =========================
        with stab2:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">📊 Classement — ' + cls["nom"] + '</div>', unsafe_allow_html=True)

            etudiants = cls["etudiants"]
            sorted_etudiants = sorted(etudiants.items(), key=lambda x: x[1]["score"], reverse=True)

            rows_html = ""
            for rang, (eid, e) in enumerate(sorted_etudiants, 1):
                score = e["score"]
                mention, m_cls = get_mention(score)
                if rang == 1:   medal = "🥇"
                elif rang == 2: medal = "🥈"
                elif rang == 3: medal = "🥉"
                else:           medal = str(rang)
                bg = "#E8F5E9" if score >= 14 else ("#E0F2F1" if score >= 10 else "#FFEBEE")
                sc_color = "#2E7D32" if score > 0 else ("#C62828" if score < 0 else "#546E6A")
                rows_html += (
                    '<tr style="background:' + bg + '">'
                    '<td style="text-align:center;font-weight:800;font-size:15px">' + str(medal) + '</td>'
                    '<td style="font-weight:700;padding-left:8px">' + e["nom"] + '</td>'
                    '<td style="text-align:center;font-weight:900;font-size:20px;color:' + sc_color + '">' + str(score) + '</td>'
                    '<td style="text-align:center"><span style="background:white;border-radius:12px;'
                    'padding:2px 10px;font-size:11px;font-weight:700">' + mention + '</span></td>'
                    '<td style="text-align:center;color:#546E6A;font-size:12px">'
                    + str(len(e["historique"])) + ' opérations</td>'
                    '</tr>'
                )

            table_html = (
                '<table style="width:100%;border-collapse:collapse;font-size:13px">'
                '<thead><tr style="background:#004D40;color:white">'
                '<th style="padding:10px 6px">Rang</th>'
                '<th style="padding:10px 8px;text-align:left">Étudiant</th>'
                '<th style="padding:10px 6px">Score</th>'
                '<th style="padding:10px 6px">Mention</th>'
                '<th style="padding:10px 6px">Historique</th>'
                '</tr></thead>'
                '<tbody>' + rows_html + '</tbody>'
                '</table>'
            )
            st.markdown(table_html, unsafe_allow_html=True)
            st.markdown("---")

            # Export Excel et PDF (identique à l'original, non modifié)
            dl1, dl2 = st.columns(2)
            with dl1:
                st.markdown('<div class="dl-excel">', unsafe_allow_html=True)
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Classement"
                thin = Side(style="thin", color="C8D8D5")
                bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
                teal_fill = PatternFill("solid", fgColor="004D40")
                mint_fill = PatternFill("solid", fgColor="E0F2F1")

                ws.merge_cells("A1:E1")
                c = ws.cell(row=1, column=1, value=cls["nom"] + " — " + cls["annee"])
                c.fill = teal_fill; c.font = Font(bold=True, color="FFFFFF", size=13)
                c.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30

                for col, h in enumerate(["Rang","Nom Étudiant","Score","Mention","Nb Opérations"], 1):
                    x = ws.cell(row=2, column=col, value=h)
                    x.fill = PatternFill("solid", fgColor="00796B")
                    x.font = Font(bold=True, color="FFFFFF", size=10)
                    x.alignment = Alignment(horizontal="center", vertical="center")
                    x.border = bdr

                for r, (eid, e) in enumerate(sorted_etudiants, 3):
                    mention, _ = get_mention(e["score"])
                    fill = mint_fill if r % 2 == 0 else None
                    for col, val in enumerate([r-2, e["nom"], e["score"], mention, len(e["historique"])], 1):
                        x = ws.cell(row=r, column=col, value=val)
                        if fill: x.fill = fill
                        x.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
                        x.border = bdr
                        if col == 2: x.font = Font(bold=True)

                for col, width in zip("ABCDE", [8, 30, 10, 15, 16]):
                    ws.column_dimensions[col].width = width

                buf = io.BytesIO()
                wb.save(buf); buf.seek(0)
                st.download_button("📊 Exporter Excel",
                    data=buf,
                    file_name=cls["nom"] + "_classement.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_cls_excel")
                st.markdown('</div>', unsafe_allow_html=True)

            with dl2:
                st.markdown('<div class="dl-pdf">', unsafe_allow_html=True)
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib import colors as rl_colors
                    from reportlab.lib.units import cm
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont
                    import os as _os

                    pdf_buf = io.BytesIO()
                    doc = SimpleDocTemplate(pdf_buf, pagesize=A4,
                                            leftMargin=2*cm, rightMargin=2*cm,
                                            topMargin=2*cm, bottomMargin=2*cm)
                    styles = getSampleStyleSheet()
                    TEAL  = rl_colors.HexColor("#004D40")
                    TEAL2 = rl_colors.HexColor("#00796B")
                    MINT  = rl_colors.HexColor("#E0F2F1")
                    GOLD  = rl_colors.HexColor("#FFF8E1")
                    GREEN = rl_colors.HexColor("#E8F5E9")
                    RED   = rl_colors.HexColor("#FFEBEE")

                    ar_font = "Helvetica"
                    try:
                        script_dir = _os.path.dirname(_os.path.abspath(__file__)) if "__file__" in dir() else "/mount/src/gestion-classe"
                        amiri = _os.path.join(script_dir, "Amiri-Regular.ttf")
                        if not _os.path.exists(amiri): amiri = "Amiri-Regular.ttf"
                        if _os.path.exists(amiri):
                            pdfmetrics.registerFont(TTFont("Amiri", amiri))
                            ar_font = "Amiri"
                    except Exception:
                        pass

                    def ar_pdf(text):
                        try:
                            import arabic_reshaper
                            from bidi.algorithm import get_display
                            return get_display(arabic_reshaper.reshape(str(text)))
                        except Exception:
                            t = str(text).strip()
                            if any('\u0600' <= c <= '\u06FF' for c in t):
                                return ' '.join(reversed(t.split()))
                            return t

                    title_s = ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold", textColor=rl_colors.white)
                    stat_s  = ParagraphStyle("st", fontSize=10, fontName="Helvetica-Bold", textColor=TEAL, spaceAfter=4)
                    story   = []

                    ht = Table([[Paragraph(cls["nom"] + " — " + cls["annee"], title_s)]],
                               colWidths=[17*cm])
                    ht.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(-1,-1), TEAL),
                        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                        ("LEFTPADDING",(0,0),(-1,-1),16),
                        ("TOPPADDING",(0,0),(-1,-1),14),
                        ("BOTTOMPADDING",(0,0),(-1,-1),14),
                    ]))
                    story.append(ht)
                    story.append(Spacer(1, 0.4*cm))

                    scores_all = [e["score"] for e in etudiants.values()]
                    moy_all = sum(scores_all)/len(scores_all) if scores_all else 0
                    story.append(Paragraph(
                        "<b>Effectif :</b> " + str(len(etudiants)) +
                        "  |  <b>Moyenne :</b> " + "{:.2f}".format(moy_all) +
                        "  |  <b>Génération :</b> " + datetime.now().strftime("%d/%m/%Y"), stat_s))
                    story.append(Spacer(1, 0.3*cm))

                    tdata = [["Rang","Nom Étudiant","Score","Mention","Opérations"]]
                    for rang, (eid, e) in enumerate(sorted_etudiants, 1):
                        mention, _ = get_mention(e["score"])
                        tdata.append([str(rang), ar_pdf(e["nom"]), str(e["score"]), mention, str(len(e["historique"]))])
                    tdata.append(["—","MOYENNE CLASSE","{:.2f}".format(moy_all),"—","—"])

                    t = Table(tdata, colWidths=[1.5*cm, 7*cm, 2.5*cm, 3*cm, 3*cm], repeatRows=1)
                    ts_list = [
                        ("BACKGROUND",    (0,0),(-1,0),  TEAL),
                        ("TEXTCOLOR",     (0,0),(-1,0),  rl_colors.white),
                        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
                        ("FONTSIZE",      (0,0),(-1,-1), 9),
                        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
                        ("FONTNAME",      (1,1),(1,-2),  ar_font),
                        ("ALIGN",         (1,1),(1,-2),  "RIGHT"),
                        ("GRID",          (0,0),(-1,-1), 0.4, rl_colors.HexColor("#C8D8D5")),
                        ("TOPPADDING",    (0,0),(-1,-1), 6),
                        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
                        ("BACKGROUND",    (0,-1),(-1,-1), GOLD),
                        ("FONTNAME",      (0,-1),(-1,-1), "Helvetica-Bold"),
                    ]
                    for rang, (eid, e) in enumerate(sorted_etudiants, 1):
                        fill = GREEN if e["score"] >= 10 else RED
                        ts_list.append(("BACKGROUND", (0,rang), (-1,rang), fill))
                    t.setStyle(TableStyle(ts_list))
                    story.append(t)
                    doc.build(story)
                    pdf_buf.seek(0)

                    st.download_button("📄 Exporter PDF",
                        data=pdf_buf,
                        file_name=cls["nom"] + "_classement.pdf",
                        mime="application/pdf",
                        key="dl_cls_pdf")
                except ImportError:
                    st.info("Ajoutez `reportlab` à requirements.txt")
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

        with stab3:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">📜 Historique des notes</div>', unsafe_allow_html=True)

            etudiants = cls["etudiants"]
            etudiant_names = {eid: e["nom"] for eid, e in etudiants.items()}
            selected_etud = st.selectbox("Choisir un étudiant",
                                          options=["Tous"] + list(etudiant_names.keys()),
                                          format_func=lambda x: "Tous les étudiants" if x == "Tous" else etudiant_names[x])

            if selected_etud == "Tous":
                all_hist = []
                for eid, e in etudiants.items():
                    for h in e["historique"]:
                        all_hist.append({
                            "Date":     h.get("date","—"),
                            "Étudiant": e["nom"],
                            "Note":     h.get("note", 0),
                            "Motif":    h.get("motif","—") or "—",
                            "Type":     "Global" if h.get("type")=="global" else "Individuel"
                        })
                if all_hist:
                    df_hist = pd.DataFrame(all_hist).sort_values("Date", ascending=False)
                    st.dataframe(df_hist, use_container_width=True, hide_index=True)
                else:
                    st.info("Aucune note saisie pour l'instant.")
            else:
                e = etudiants[selected_etud]
                st.markdown(
                    '<b style="font-size:16px">' + e["nom"] + '</b>'
                    ' &nbsp;— Score total : <b style="color:#00796B;font-size:18px">'
                    + str(e["score"]) + '</b>',
                    unsafe_allow_html=True
                )
                if e["historique"]:
                    df_h = pd.DataFrame(e["historique"]).rename(columns={
                        "note":"Note","motif":"Motif","date":"Date","type":"Type"})
                    df_h["Type"] = df_h["Type"].map({"global":"Global","indiv":"Individuel"}).fillna("—")
                    df_h = df_h[["Date","Note","Motif","Type"]].sort_values("Date", ascending=False)
                    st.dataframe(df_h, use_container_width=True, hide_index=True)
                    cumul = []
                    running = 0
                    for h in reversed(e["historique"]):
                        running += h.get("note", 0)
                        cumul.append(running)
                    st.line_chart(pd.DataFrame({"Score cumulé": cumul}), height=200)
                else:
                    st.info("Aucune note pour cet étudiant.")

            st.markdown('</div>', unsafe_allow_html=True)

        with stab4:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">⚙️ Gestion de la classe — ' + cls["nom"] + '</div>', unsafe_allow_html=True)

            etudiants = cls["etudiants"]

            st.markdown("**✏️ Modifier un étudiant**")
            eid_to_edit = st.selectbox("Étudiant à modifier",
                                        options=list(etudiants.keys()),
                                        format_func=lambda x: etudiants[x]["nom"],
                                        key="edit_select")
            col_edit1, col_edit2, col_edit3 = st.columns([3,2,1])
            with col_edit1:
                new_nom = st.text_input("Nouveau nom", value=etudiants[eid_to_edit]["nom"], key="new_nom")
            with col_edit2:
                new_score = st.number_input("Réinitialiser le score", value=etudiants[eid_to_edit]["score"], step=1, key="new_score")
            with col_edit3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Sauver", key="save_edit"):
                    etudiants[eid_to_edit]["nom"]   = new_nom.strip()
                    etudiants[eid_to_edit]["score"] = int(new_score)
                    save_db(db)
                    st.success("✅ Modifié !")
                    st.rerun()

            st.markdown("---")

            st.markdown("**➕ Ajouter un étudiant**")
            col_add1, col_add2 = st.columns([4,1])
            with col_add1:
                new_student = st.text_input("Nom du nouvel étudiant", key="add_student")
            with col_add2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("➕ Ajouter", key="btn_add"):
                    if new_student.strip():
                        new_id = str(max([int(k) for k in etudiants.keys()], default=-1) + 1)
                        etudiants[new_id] = {"nom": new_student.strip(), "score": 0, "historique": []}
                        save_db(db)
                        st.success("✅ Étudiant ajouté !")
                        st.rerun()

            st.markdown("---")

            st.markdown("**🗑️ Supprimer un étudiant**")
            eid_to_del = st.selectbox("Étudiant à supprimer",
                                       options=list(etudiants.keys()),
                                       format_func=lambda x: etudiants[x]["nom"],
                                       key="del_select")
            if st.button("🗑️ Supprimer " + etudiants[eid_to_del]["nom"], key="btn_del"):
                st.session_state.confirm_delete = eid_to_del

            if st.session_state.confirm_delete == eid_to_del and eid_to_del in etudiants:
                st.warning("⚠️ Confirmer la suppression de **" + etudiants[eid_to_del]["nom"] + "** ?")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("✅ Confirmer", key="confirm_del"):
                        del etudiants[eid_to_del]
                        save_db(db)
                        st.session_state.confirm_delete = None
                        st.success("Étudiant supprimé.")
                        st.rerun()
                with c2:
                    if st.button("❌ Annuler", key="cancel_del"):
                        st.session_state.confirm_delete = None
                        st.rerun()

            st.markdown("---")

            st.markdown("**⛔ Supprimer toute la classe**")
            if st.button("⛔ Supprimer la classe " + cls["nom"], key="del_class"):
                st.session_state.confirm_delete = "CLASS_" + selected
            if st.session_state.confirm_delete == "CLASS_" + selected:
                st.error("⚠️ Cette action est irréversible ! Supprimer la classe **" + cls["nom"] + "** ?")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("✅ Oui, supprimer", key="confirm_del_class"):
                        del db[selected]
                        save_db(db)
                        st.session_state.selected_class = None
                        st.session_state.confirm_delete = None
                        st.success("Classe supprimée.")
                        st.rerun()
                with c2:
                    if st.button("❌ Annuler", key="cancel_del_class"):
                        st.session_state.confirm_delete = None
                        st.rerun()

            st.markdown('</div>', unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="app-footer">
    <span class="badge">Gestion Classe v1.0</span> &nbsp;|&nbsp;
    Suivi des notes de comportement &nbsp;|&nbsp;
    {datetime.now().strftime('%Y')}
</div>
""", unsafe_allow_html=True)
